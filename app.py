import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import threading
from datetime import datetime
from Levenshtein import distance as levenshtein_distance
import openpyxl

class AddressSimilarityApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # 设置窗口属性
        self.title("光交地址相似度匹配工具")
        self.geometry("800x600")

        # 设置主题和颜色
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        # 初始化变量
        self.selected_file = ""
        self.column_options = []
        self.selected_col1 = tk.StringVar()
        self.selected_col2 = tk.StringVar()

        # 关键词过滤列表（从批量处理.py复制）
        self.address_keywords_to_remove = ['光交', ':', '-', '号', '(', ')', '（', '）', ' ']

        self.setup_ui()

    def setup_ui(self):
        """设置用户界面"""
        # 主框架
        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # 标题
        title_label = ctk.CTkLabel(
            self.main_frame,
            text="光交地址相似度匹配工具",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=20)

        # 文件选择区域
        file_frame = ctk.CTkFrame(self.main_frame)
        file_frame.pack(fill="x", padx=20, pady=10)

        file_label = ctk.CTkLabel(file_frame, text="选择文件：", font=ctk.CTkFont(size=14))
        file_label.pack(side="left", padx=(0, 10))

        self.file_path_label = ctk.CTkLabel(
            file_frame,
            text="未选择文件",
            font=ctk.CTkFont(size=12),
            fg_color="transparent"
        )
        self.file_path_label.pack(side="left", fill="x", expand=True)

        select_file_btn = ctk.CTkButton(
            file_frame,
            text="选择文件",
            command=self.select_file,
            width=100
        )
        select_file_btn.pack(side="right")

        # 列选择区域
        column_frame = ctk.CTkFrame(self.main_frame)
        column_frame.pack(fill="x", padx=20, pady=10)

        # 第一列选择
        col1_frame = ctk.CTkFrame(column_frame)
        col1_frame.pack(fill="x", pady=5)

        col1_label = ctk.CTkLabel(col1_frame, text="选择对比列1：", font=ctk.CTkFont(size=14))
        col1_label.pack(side="left", padx=(0, 10))

        self.col1_combobox = ctk.CTkComboBox(
            col1_frame,
            variable=self.selected_col1,
            values=["请选择文件后加载"],
            state="readonly",
            width=300
        )
        self.col1_combobox.pack(side="left", fill="x", expand=True)

        # 第二列选择
        col2_frame = ctk.CTkFrame(column_frame)
        col2_frame.pack(fill="x", pady=5)

        col2_label = ctk.CTkLabel(col2_frame, text="选择对比列2：", font=ctk.CTkFont(size=14))
        col2_label.pack(side="left", padx=(0, 10))

        self.col2_combobox = ctk.CTkComboBox(
            col2_frame,
            variable=self.selected_col2,
            values=["请选择文件后加载"],
            state="readonly",
            width=300
        )
        self.col2_combobox.pack(side="left", fill="x", expand=True)

        # 运行按钮
        self.run_button = ctk.CTkButton(
            self.main_frame,
            text="运行匹配",
            command=self.run_similarity_matching,
            font=ctk.CTkFont(size=16, weight="bold"),
            height=40
        )
        self.run_button.pack(pady=20)

        # 进度显示区域
        log_frame = ctk.CTkFrame(self.main_frame)
        log_frame.pack(fill="both", expand=True, padx=20, pady=10)

        log_label = ctk.CTkLabel(log_frame, text="执行日志：", font=ctk.CTkFont(size=14))
        log_label.pack(anchor="w", pady=(0, 5))

        # 创建文本区域和滚动条
        self.log_text = tk.Text(
            log_frame,
            height=15,
            wrap="word",
            font=("Consolas", 10)
        )

        scrollbar = ctk.CTkScrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        self.log_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 设置文本区域为只读
        self.log_text.configure(state="disabled")

        # 添加一些初始提示信息
        self.log_message("欢迎使用光交地址相似度匹配工具！")
        self.log_message("请先选择Excel文件，然后选择要对比的两列，最后点击运行匹配。")

    def log_message(self, message):
        """添加日志消息"""
        self.log_text.configure(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")  # 滚动到底部
        self.log_text.configure(state="disabled")
        self.update_idletasks()  # 立即更新界面

    def select_file(self):
        """选择Excel文件"""
        filetypes = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]

        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=filetypes
        )

        if filename:
            self.selected_file = filename
            self.file_path_label.configure(text=filename)

            # 读取文件列名
            try:
                self.load_column_names(filename)
                self.log_message(f"成功加载文件：{os.path.basename(filename)}")
            except Exception as e:
                self.log_message(f"读取文件时出错：{str(e)}")
                messagebox.showerror("错误", f"读取文件时出错：{str(e)}")

    def load_column_names(self, filename):
        """加载Excel文件的列名"""
        try:
            workbook = openpyxl.load_workbook(filename, read_only=True)
            sheet = workbook.active
            self.column_options = [cell.value for cell in sheet[1]]
            workbook.close()

            # 更新下拉框选项
            self.col1_combobox.configure(values=self.column_options)
            self.col2_combobox.configure(values=self.column_options)

            # 设置默认选择（如果有对应的列）
            default_cols = ['最早光交名称', '竣工光交名称']
            if default_cols[0] in self.column_options:
                self.selected_col1.set(default_cols[0])
            if default_cols[1] in self.column_options:
                self.selected_col2.set(default_cols[1])

            self.log_message(f"文件包含 {len(self.column_options)} 列：{', '.join(self.column_options[:5])}...")

        except Exception as e:
            raise Exception(f"使用openpyxl读取列名失败：{str(e)}")

    def run_similarity_matching(self):
        """运行相似度匹配"""
        # 验证输入
        if not self.selected_file:
            messagebox.showerror("错误", "请先选择Excel文件！")
            return

        col1 = self.selected_col1.get()
        col2 = self.selected_col2.get()

        if not col1 or not col2:
            messagebox.showerror("错误", "请选择要对比的两列！")
            return

        if col1 == col2:
            messagebox.showerror("错误", "请选择不同的列进行对比！")
            return

        # 禁用运行按钮
        self.run_button.configure(state="disabled", text="处理中...")

        # 在新线程中运行，避免界面冻结
        thread = threading.Thread(target=self.process_file, args=(col1, col2))
        thread.daemon = True
        thread.start()

    def process_file(self, col1, col2):
        """处理文件（在后台线程中运行）"""
        try:
            self.log_message("开始处理文件...")
            self.log_message(f"对比列：{col1} vs {col2}")

            # 读取数据
            self.log_message("正在读取Excel文件 (使用 openpyxl)...")
            workbook = openpyxl.load_workbook(self.selected_file)
            sheet = workbook.active

            # 获取表头和列索引
            headers = [cell.value for cell in sheet[1]]
            if col1 not in headers or col2 not in headers:
                raise ValueError(f"选择的列在文件中不存在！")
            
            col1_idx = headers.index(col1) + 1
            col2_idx = headers.index(col2) + 1

            # 添加新的“相似度”列
            similarity_col_idx = sheet.max_column + 1
            sheet.cell(row=1, column=similarity_col_idx, value="相似度")

            # 计算相似度
            self.log_message(f"开始计算地址相似度 (共 {sheet.max_row - 1} 行)...")
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                addr1 = row[col1_idx - 1]
                addr2 = row[col2_idx - 1]
                similarity = self.calculate_address_similarity(addr1, addr2)
                sheet.cell(row=i + 2, column=similarity_col_idx, value=similarity)
                if (i + 1) % 100 == 0:
                    self.log_message(f"已处理 {i + 1} 行...")

            # 生成输出文件名
            input_filename = os.path.basename(self.selected_file)
            output_filename = f"【已匹配】{input_filename}"
            output_path = os.path.join(os.path.dirname(self.selected_file), output_filename)

            # 保存结果
            self.log_message(f"正在保存结果至：{output_filename}")
            workbook.save(output_path)
            workbook.close()

            # 完成
            self.log_message(f"处理完成！共处理 {sheet.max_row - 1} 行。")
            self.log_message(f"输出文件：{output_path}")

            # 在主线程中显示成功消息
            self.after(0, lambda: messagebox.showinfo(
                "成功",
                f"处理完成！\n输出文件：{output_filename}"
            ))

        except Exception as e:
            error_msg = f"处理过程中出错：{str(e)}"
            self.log_message(error_msg)

            # 在主线程中显示错误消息
            self.after(0, lambda: messagebox.showerror("错误", error_msg))

        finally:
            # 重新启用运行按钮
            self.after(0, lambda: self.run_button.configure(state="normal", text="运行匹配"))

    def calculate_address_similarity(self, addr1, addr2):
        """计算地址相似度（从批量处理.py复制的核心算法）"""
        # 设置默认的目标关键词列表
        target_prefixes = [
            "光交", "OBD", "分光器", "小区", "机房", "FTTR", "FTTH", "单元", "幛",
            "接入点", "扬州", "邗江", "广陵", "江都", "高邮", "宝应", "仪征",
            "三网", "开发商", "ODF", "用户", "接入网", "无线"
        ]

        # 统一处理空值和格式，并过滤关键词
        addr1_processed = self.filter_keywords(
            str(addr1).strip() if addr1 is not None else "",
            target_prefixes
        )
        addr2_processed = self.filter_keywords(
            str(addr2).strip() if addr2 is not None else "",
            target_prefixes
        )

        # 规则1：双向检测过滤后的子字符串包含关系
        if addr2_processed in addr1_processed or addr1_processed in addr2_processed:
            return round(1.0, 2)

        # 规则2：Levenshtein编辑距离计算相似度
        max_len = max(len(addr1_processed), len(addr2_processed))
        if max_len == 0:  # 两个过滤后的地址均为空
            return round(1.0, 2)

        distance = levenshtein_distance(addr1_processed, addr2_processed)
        similarity = 1 - (distance / max_len)
        return round(similarity, 2)

    def filter_keywords(self, input_str, target_prefixes):
        """过滤关键词（从批量处理.py复制）"""
        # 对关键词按长度降序排序
        sorted_prefixes = sorted(target_prefixes, key=lambda x: len(x), reverse=True)

        while True:
            original_str = input_str
            # 遍历所有关键词进行替换
            for prefix in sorted_prefixes:
                input_str = input_str.replace(prefix, '')

            # 如果本轮没有发生任何替换，退出循环
            if input_str == original_str:
                break

        return input_str

if __name__ == "__main__":
    app = AddressSimilarityApp()
    app.mainloop()